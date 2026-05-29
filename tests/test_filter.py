"""Tests for _mojo/filter.py — county-based row filtering."""

import logging
from pathlib import Path

import openpyxl

from _mojo.filter import COUNTY_COLUMN, filter_by_county


def _make_xlsx(tmp_path: Path, rows: list[list]) -> Path:
    """Create a minimal XLSX with a header row followed by *rows*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Source", COUNTY_COLUMN, "Phone 1"])
    for row in rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def _read_counties(path: Path) -> list[str | None]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return [row[1].value for row in ws.iter_rows(min_row=2)]


def test_filter_keeps_matching_rows(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["src", "Chester", "555-0001"],
        ["src", "Other County", "555-0002"],
        ["src", "Philadelphia", "555-0003"],
    ])
    assert filter_by_county(path, {"Chester", "Philadelphia"}, "FSBO") is True
    assert _read_counties(path) == ["Chester", "Philadelphia"]


def test_filter_case_insensitive(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["src", "new castle", "555-0001"],
        ["src", "CHESTER", "555-0002"],
        ["src", "Other", "555-0003"],
    ])
    assert filter_by_county(path, {"New Castle", "Chester"}, "FSBO") is True
    assert _read_counties(path) == ["new castle", "CHESTER"]


def test_filter_no_op_empty_counties(tmp_path):
    path = _make_xlsx(tmp_path, [
        ["src", "Chester", "555-0001"],
        ["src", "Other County", "555-0002"],
    ])
    mtime_before = path.stat().st_mtime
    assert filter_by_county(path, set(), "FSBO") is True
    assert path.stat().st_mtime == mtime_before  # file untouched


def test_filter_missing_column_warns(tmp_path, caplog):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Source", "NotCountyColumn", "Phone 1"])
    ws.append(["src", "Chester", "555-0001"])
    path = tmp_path / "no_county_col.xlsx"
    wb.save(path)

    with caplog.at_level(logging.WARNING, logger="mojo_downloader"):
        result = filter_by_county(path, {"Chester"}, "FSBO")

    assert result is False
    assert "column not found" in caplog.text
    # File should be unchanged — still has the original non-matching header
    wb2 = openpyxl.load_workbook(path)
    assert wb2.active.max_row == 2
