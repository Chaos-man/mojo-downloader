"""Post-download XLSX city filter."""

import logging
from pathlib import Path

import openpyxl

log = logging.getLogger("mojo_downloader")

COUNTY_COLUMN = "County"


def filter_by_county(path: Path, counties: set[str], label: str) -> bool:
    """Remove rows from the XLSX at *path* whose county is not in *counties*.

    Overwrites the file in place. Logs kept/total counts on one line.
    Logs a warning (not an error) if the county column is absent.
    No-op if *counties* is empty.

    Returns True if filtering ran (or was skipped because counties is empty).
    Returns False if the county column was not found — caller should notify.
    """
    if not counties:
        return True

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Locate the county column by header name (case-insensitive).
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        county_col = next(
            i for i, h in enumerate(headers)
            if h and h.strip().lower() == COUNTY_COLUMN.lower()
        )
    except StopIteration:
        log.warning("%s: '%s' column not found — county filter skipped", label, COUNTY_COLUMN)
        return False

    counties_lower = {c.strip().lower() for c in counties}
    rows_to_delete = []
    total = ws.max_row - 1  # exclude header

    for row in ws.iter_rows(min_row=2):
        val = row[county_col].value
        if val is None or val.strip().lower() not in counties_lower:
            rows_to_delete.append(row[0].row)

    # Delete in reverse order so row numbers stay valid.
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)

    kept = total - len(rows_to_delete)
    log.info(
        "%s: kept %d/%d rows (%d removed by county filter)",
        label, kept, total, len(rows_to_delete),
    )
    wb.save(path)
    return True
